import cv2
import numpy as np
import torch
import pandas as pd
import re
import os
import random
from torch.utils.data import Dataset
import random
import math
import csv
import pickle
from openpyxl import load_workbook

class SegDataSet(Dataset):
    def __init__(self, base_dir, batch_size, is_train=True):
        super().__init__()

        self.isTrain = is_train
        self.batch_size = batch_size

        self.all_data = []
        pkl_path = 'tensor.pkl'
        if os.path.exists(pkl_path):
            with open(pkl_path, 'rb') as fp:
                self.all_data = pickle.load(fp)
        else:
            tmp_dict = {}
            for dirpath, dirnames, filenames in os.walk(base_dir, followlinks=True):
                # Es1
                pattern = re.compile(r'^[^\.].*\.(csv|xlsx)$')
                for filename in filenames:
                    if not pattern.match(filename): continue
                    if "Es1" not in dirpath:
                        continue
                    if 'JointPosition_new' in filename:
                        position_path = os.path.join(dirpath, filename)
                        label_dir_path = dirpath.replace('Raw','label')
                        for dirpath, dirnames, filename in os.walk(label_dir_path):
                            for file in filename:
                                if 'Clinical' in file:
                                    label_path = os.path.join(label_dir_path, file)
                                    if os.path.exists(position_path) and os.path.exists(label_path):
                                        # read label
                                        wb = load_workbook(label_path)
                                        sheets = wb.worksheets
                                        sheet1 = sheets[0]
                                        label = sheet1.cell(2, 2).value
                                        # read data
                                        data =pd.read_csv(position_path,index_col=None,usecols=None)
                                        #f = open(position_path, 'r', encoding='utf-8')
                                        tensor_data = torch.tensor(data.iloc[0])
                                        tensor_data=tensor_data.unsqueeze(dim=0)

                                        for i in range(1,len(data) - 104):
                                            position_data = []
                                            print('======>load {}/{}:'.format(position_path,i))
                                            for j in range(i, i+104, 2):
                                                darray = torch.tensor(data.iloc[j]).unsqueeze(dim=0)
                                                tensor_data = torch.cat((tensor_data, darray),0)
                                            i += 5
                                            # split test and train
                                            ID_num = int(dirpath.split("/")[-3].split("ID")[1])
                                            if ID_num > 3:
                                                tmp_dict['train_data'] = tensor_data
                                                tmp_dict['train_label'] = label
                                                self.all_data.append(tmp_dict)
                                            else:
                                                continue

        with open(pkl_path, 'wb') as fp:
            pickle.dump(self.all_data, fp)

        print('All Data Len:{}'.format(len(self.all_data)))

    def __len__(self):
        return len(self.all_data)

    def __getitem__(self, index):


        data = self.all_data[index]['train_data']
        label = self.all_data[index]['train_label']

        ret_dic = {'train_data': data,
                   'train_label':label,
                  }
        return ret_dic

if __name__=='__main__':
    from options import BaseOptions
    import dataloader
    import torch
    opt = BaseOptions().parse()
    opt.isTrain = True

    trainset = dataloader.SegDataSet(opt.train_data_dir, opt.train_batch_size, is_train=opt.isTrain)
    trainloader = torch.utils.data.DataLoader(trainset,
                                              batch_size=4)
    for idx, batch in enumerate(trainloader):
        a = batch['train_data']
        print(a.size())

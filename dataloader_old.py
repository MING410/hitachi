import cv2
import numpy as np
import torch
import pandas as pd
import re
import os
import random
from torch.utils.data import Dataset
import random
import math
import csv
import pickle
from openpyxl import load_workbook

class SegDataSet(Dataset):
    def __init__(self, base_dir, batch_size, is_train=True):
        super().__init__()

        self.isTrain = is_train
        self.batch_size = batch_size

        self.all_data = []
        pkl_path = 't.pkl'
        if os.path.exists(pkl_path):
            with open(pkl_path, 'rb') as fp:
                self.all_data = pickle.load(fp)
        else:
            tmp_dict = {}
            for dirpath, dirnames, filenames in os.walk(base_dir, followlinks=True):
                # Es1
                pattern = re.compile(r'^[^\.].*\.(csv|xlsx)$')
                for filename in filenames:
                    if not pattern.match(filename): continue
                    if "Es1" not in dirpath:
                        continue
                    if 'JointPosition_new_new' in filename:
                        position_path = os.path.join(dirpath, filename)
                        label_dir_path = dirpath.replace('Raw','label')
                        for dirpath, dirnames, filename in os.walk(label_dir_path):
                            for file in filename:
                                if 'Clinical' in file:
                                    label_path = os.path.join(label_dir_path, file)
                                    if os.path.exists(position_path) and os.path.exists(label_path):
                                        # read label
                                        wb = load_workbook(label_path)
                                        sheets = wb.worksheets
                                        sheet1 = sheets[0]
                                        label = sheet1.cell(2, 2).value
                                        # read data
                                        f = open(position_path, 'r', encoding='utf-8')
                                        file_list=[]
                                        for line in f:
                                            int_line=[]
                                            for i in line:
                                                if len(i)<3:
                                                    continue
                                                int_line.append(float(i.strip()))
                                            if len(int_line) != 0:
                                                file_list.append(int_line)
                                        # 27 lines a data
                                        for i in range(len(file_list) - 54):
                                            position_data = []
                                            for j in range(i, i+54, 2):
                                                darray = file_list[j]
                                                position_data.append(darray)
                                            i += 3
                                            # split test and train
                                            ID_num = int(dirpath.split("/")[-3].split("ID")[1])
                                            if ID_num > 3:
                                                tmp_dict['train_data'] = position_data
                                                tmp_dict['train_label'] = label
                                                self.all_data.append(tmp_dict)
                                            else:
                                                continue



        with open(pkl_path, 'wb') as fp:
            pickle.dump(self.all_data, fp)

        print('All Data Len:{}'.format(len(self.all_data)))

    def __len__(self):
        return len(self.all_data)

    def __getitem__(self, index):


        data = self.all_data[index]['train_data']
        label = self.all_data[index]['train_label']

        ret_dic = {'train_data': data,
                   'train_label':label,
                  }
        return ret_dic

if __name__=='__main__':
    from options import BaseOptions
    import dataloader_old
    import torch
    opt = BaseOptions().parse()
    opt.isTrain = True

    trainset = dataloader_old.SegDataSet(opt.train_data_dir, opt.train_batch_size, is_train=opt.isTrain)
    trainloader = torch.utils.data.DataLoader(trainset,
                                              batch_size=4)
    # for idx, batch in enumerate(trainloader):
    #     a = np.array(batch['train_data'])
    #     print(a.size())
